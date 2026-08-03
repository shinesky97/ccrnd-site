# -*- coding: utf-8 -*-
"""엑셀 조작 계획(Operation Plan)과 실행 엔진.

모든 엑셀 변경은 '계획(op 목록)'으로 먼저 만들어 Dry-run으로 보여준 뒤 실행한다.
실행 엔진 2종:
- XlwingsEngine : Windows + Excel 설치 환경 권장. 실제 Excel로 조작하므로
  수식 참조·서식·차트·도형이 완전 보존된다 (열 삽입 시 참조 자동 조정 포함).
- OpenpyxlEngine: Excel 없는 환경(개발·검증용). 값/텍스트 변경은 안전하지만
  열 삽입 시 수식 참조가 조정되지 않으므로 insert_col 이 포함된 계획은 거부한다.
"""
import os
import shutil

# ---- 연산 정의 (파일 단위 계획) ----
# op 종류:
#  {'op':'set_cell','sheet':s,'cell':'B3','value':v,'old':...}
#  {'op':'clear_cell','sheet':s,'cell':'B3','old':...}
#  {'op':'copy_col_values','sheet':s,'src_col':3,'dst_col':4,'row_from':r1,'row_to':r2}
#  {'op':'insert_col','sheet':s,'at':5}   # xlwings 전용
#  {'op':'note','text':...}               # 실행 없음(리포트 표시용)


class Plan:
    def __init__(self, src, out):
        self.src = src          # 원본 (읽기 전용)
        self.out = out          # 산출물 경로 (신규 생성)
        self.ops = []
        self.warnings = []
        self.manual = []        # 자동화 불가 — 수동 처리 목록

    def add(self, **op):
        self.ops.append(op)

    def describe(self):
        lines = [f'원본: {os.path.basename(self.src)}',
                 f'생성: {os.path.basename(self.out)}',
                 f'작업 {len(self.ops)}건:']
        from collections import Counter
        c = Counter(o['op'] for o in self.ops)
        for k, n in c.items():
            lines.append(f'  - {k}: {n}건')
        for w in self.warnings:
            lines.append(f'  ⚠ {w}')
        for m in self.manual:
            lines.append(f'  ✋ 수동: {m}')
        return '\n'.join(lines)


class OpenpyxlEngine:
    name = 'openpyxl'

    def execute(self, plan):
        if any(o['op'] == 'insert_col' for o in plan.ops):
            raise RuntimeError(
                'openpyxl 엔진은 열 삽입 시 수식 참조를 조정하지 못해 이 계획을 실행할 수 '
                '없습니다. Windows에서 xlwings 엔진(Excel)으로 실행하십시오.')
        from openpyxl import load_workbook
        from .util import patch_openpyxl_korean
        patch_openpyxl_korean()
        shutil.copy2(plan.src, plan.out)
        wb = load_workbook(plan.out)          # 수식 보존 로드
        vwb = None                            # 값 조회용 (copy_col_values)
        applied = 0
        for o in plan.ops:
            if o['op'] == 'note':
                continue
            ws = wb[o['sheet']]
            if o['op'] in ('set_cell', 'clear_cell'):
                cell = ws[o['cell']]
                try:
                    cell.value = o.get('value') if o['op'] == 'set_cell' else None
                    applied += 1
                except AttributeError:
                    plan.warnings.append(f"{o['sheet']}!{o['cell']} 병합셀 하위 — 건너뜀")
            elif o['op'] == 'copy_col_values':
                if vwb is None:
                    vwb = load_workbook(plan.src, data_only=True)
                vs = vwb[o['sheet']]
                for r in range(o['row_from'], o['row_to'] + 1):
                    val = vs.cell(row=r, column=o['src_col']).value
                    tgt = ws.cell(row=r, column=o['dst_col'])
                    try:
                        tgt.value = val
                        applied += 1
                    except AttributeError:
                        pass
        wb.save(plan.out)
        return applied


class XlwingsEngine:
    name = 'xlwings'

    def execute(self, plan):
        import xlwings as xw
        shutil.copy2(plan.src, plan.out)
        applied = 0
        with xw.App(visible=False, add_book=False) as app:
            app.display_alerts = False
            wb = app.books.open(os.path.abspath(plan.out))
            for o in plan.ops:
                if o['op'] == 'note':
                    continue
                sht = wb.sheets[o['sheet']]
                if o['op'] == 'set_cell':
                    sht.range(o['cell']).value = o.get('value')
                    applied += 1
                elif o['op'] == 'clear_cell':
                    sht.range(o['cell']).clear_contents()
                    applied += 1
                elif o['op'] == 'copy_col_values':
                    src = sht.range((o['row_from'], o['src_col']), (o['row_to'], o['src_col']))
                    dst = sht.range((o['row_from'], o['dst_col']))
                    dst.value = src.value  # 값만
                    applied += 1
                elif o['op'] == 'insert_col':
                    sht.api.Columns(o['at']).Insert()
                    applied += 1
            wb.save()
            wb.close()
        return applied


def pick_engine(prefer_excel=True):
    if prefer_excel:
        try:
            import xlwings  # noqa: F401
            return XlwingsEngine()
        except ImportError:
            pass
    return OpenpyxlEngine()
