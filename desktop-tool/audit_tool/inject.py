# -*- coding: utf-8 -*-
"""D3: 정산표 당기 숫자 자동 주입 + 검증 리포트.

원천: 전산자료(더존) BS/PL 시트 → 대상: 당기 정산표 AR 시트의 당기(연도) 열.

원칙:
- 매칭은 '회사제시계정과목 정규화 명칭 + 출현 순번'의 정확 일치만. 추측 매핑 금지.
- 동일 계정명의 출현 횟수가 원천과 정산표에서 다르면 해당 계정 전체를 '모호'로 제외.
- 수식이 들어 있는 셀(소계·합계)은 덮어쓰지 않는다.
- 미매칭·모호·수식보호 항목은 전부 리포트에 나열한다.
"""
import os
import re
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .engine import Plan
from .util import patch_openpyxl_korean

patch_openpyxl_korean()

SUBTOTAL_PAT = re.compile(r'^([ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.?|\(\d+\)|\d+\.)')
TOTAL_KEYS = ('자산총계', '부채총계', '자본총계', '부채와자본총계',
              '당기순이익', '당기순손실', '법인세차감전순이익')
TOTAL_ALIASES = {'부채및자본총계': '부채와자본총계'}
SECTION_WORDS = ('자산', '부채', '자본', '유동자산', '비유동자산', '유동부채', '비유동부채')
HEADER_WORDS = {'회사제시계정과목', '공시용계정과목', '과목', '계정과목', '금액'}


def _first_number(row, lo, hi):
    """row[lo:hi] 중 첫 숫자 (더존 2단 구조: 세부열/본란열 어느 쪽이든)."""
    for v in row[lo:hi]:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return v
    return None


def _norm(s):
    return re.sub(r'[\s　]+', '', str(s))


def _is_leaf(norm_name):
    if SUBTOTAL_PAT.match(norm_name):
        return False
    if norm_name in SECTION_WORDS or norm_name.endswith('총계'):
        return False
    if any(norm_name.startswith(t) for t in TOTAL_KEYS):
        return False
    return True


def load_raw_fs(path, sheets=('BS', 'PL')):
    """더존 BS/PL 파싱 → (계정명별 당기금액 리스트, 합계값, 이슈)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    amounts = defaultdict(list)   # 정규화명 → [금액, ...] (출현 순서)
    totals, issues = {}, []
    for sn in sheets:
        if sn not in wb.sheetnames:
            issues.append(f'전산자료에 {sn} 시트 없음')
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or row[0] is None or not str(row[0]).strip():
                continue
            n = _norm(row[0])
            n = TOTAL_ALIASES.get(n, n)
            b = row[1] if len(row) > 1 and isinstance(row[1], (int, float)) \
                and not isinstance(row[1], bool) else None
            c = row[2] if len(row) > 2 and isinstance(row[2], (int, float)) \
                and not isinstance(row[2], bool) else None
            if b is None and c is None:
                continue
            for tk in TOTAL_KEYS:
                if n.startswith(tk):
                    totals.setdefault(tk, b if b is not None else c)
            if n not in HEADER_WORDS:
                # 소계·총계 행도 포함 — 정산표 쪽 셀이 수식이면 자동 보호되므로 안전
                amounts[n].append((b, c))
    wb.close()
    # 원천 자체 검증: 자산총계 = 부채와자본총계 (또는 부채총계+자본총계)
    checks = []
    a = totals.get('자산총계')
    le = totals.get('부채와자본총계')
    if le is None and '부채총계' in totals and '자본총계' in totals:
        le = totals['부채총계'] + totals['자본총계']
    if a is not None and le is not None:
        ok = abs(a - le) < 1
        checks.append(('BS 대차 검증 (자산총계 = 부채와자본총계)', a, le, ok))
        if not ok:
            issues.append(f'전산자료 BS 대차 불일치: 자산총계 {a:,.0f} ≠ 부채와자본 {le:,.0f}')
    else:
        issues.append('전산자료에서 자산총계/부채와자본총계를 찾지 못함 — BS 대차 검증 불가')
    return dict(amounts), totals, checks, issues


def _find_ar_layout(ws):
    """AR 시트에서 계정명 열과 '당기 금액' 2단 입력열(F,G 형태)을 찾는다. 실패 시 None.

    입력 대상은 '제 N (당)기 금액' 헤더 아래 2개 열(세부/본란)이다.
    연도 열(=F+G 수식)과 감사후 열은 수식이므로 건드리지 않는다.
    """
    name_col = detail_col = header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            s = _norm(v)
            if s == '회사제시계정과목':
                name_col = c.column
                header_row = max(header_row or 0, c.row)
            if re.search(r'제\d+\(당\)기', s) and detail_col is None:
                detail_col = c.column
    if name_col and detail_col:
        return {'header_row': header_row, 'name_col': name_col,
                'detail_col': detail_col, 'main_col': detail_col + 1}
    return None


def _find_name_col(ws, header='회사제시계정과목', max_row=10):
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for c in row:
            if isinstance(c.value, str) and _norm(c.value) == _norm(header):
                return c.row, c.column
    return None, None


def plan_inject(trial_path, raw_path, year, sheet='AR', name_source=None):
    """주입 계획 생성. (plan, report_lines) 반환. 대상 파일은 in-place(백업 후) 수정.

    name_source: 계정명 열이 수식(=E10 등)이라 값이 비어 보일 때 이름을 보충할
    전기 정산표 경로 (행 번호 동일 전제 — 캐시값 보존 파일).
    """
    amounts, totals, checks, issues = load_raw_fs(raw_path)
    plan = Plan(trial_path, trial_path)   # in-place (실행 전 백업은 호출자가 수행)
    rpt = ['=== 정산표 당기 숫자 주입 리포트 ===',
           f'원천: {os.path.basename(raw_path)} (BS/PL)',
           f'대상: {os.path.basename(trial_path)} [{sheet}] {year}년 열', '']
    for label, a, b, ok in checks:
        rpt.append(f"{'✔' if ok else '✘'} {label}: {a:,.0f} vs {b:,.0f}")
    for i in issues:
        rpt.append(f'⚠ {i}')

    wbf = load_workbook(trial_path)               # 수식 확인용
    wbv = load_workbook(trial_path, data_only=True)
    if sheet not in wbf.sheetnames:
        plan.manual.append(f'{sheet} 시트 없음')
        rpt.append(f'✘ {sheet} 시트 없음 — 중단')
        return plan, rpt
    wsf, wsv = wbf[sheet], wbv[sheet]
    layout = _find_ar_layout(wsv)
    if layout is None:
        plan.manual.append(f'{sheet}: 회사제시계정과목/당기 금액 열 탐지 실패')
        rpt.append(f'✘ 헤더 탐지 실패: 회사제시계정과목 열과 "제N(당)기" 금액 열이 '
                   f'필요합니다. 열 제목을 확인하십시오.')
        return plan, rpt
    rpt.append(f"열 탐지: 계정명={get_column_letter(layout['name_col'])}, "
               f"당기 입력열={get_column_letter(layout['detail_col'])}(세부)/"
               f"{get_column_letter(layout['main_col'])}(본란) "
               f"(헤더 {layout['header_row']}행)")

    # 계정명 보충 소스 (수식 캐시값 소실 대응)
    src_ws = src_name_col = None
    if name_source and os.path.exists(name_source):
        src_wb = load_workbook(name_source, read_only=True, data_only=True)
        if sheet in src_wb.sheetnames:
            src_ws = src_wb[sheet]
            _, src_name_col = _find_name_col(src_ws)

    # 1차 통과: AR의 계정명 출현 횟수 집계
    ar_rows, filled_from_source = [], 0
    for r in range(layout['header_row'] + 1, wsv.max_row + 1):
        v = wsv.cell(row=r, column=layout['name_col']).value
        if (v is None or not str(v).strip()) and src_ws is not None and src_name_col:
            v = src_ws.cell(row=r, column=src_name_col).value
            if v is not None and str(v).strip():
                filled_from_source += 1
        if v is None or not str(v).strip():
            continue
        n = _norm(v)
        if n not in HEADER_WORDS and n not in SECTION_WORDS:
            ar_rows.append((r, n))
    if filled_from_source:
        rpt.append(f'계정명 {filled_from_source}행을 전기 정산표에서 보충 '
                   f'(당기 파일의 계정명 열이 수식이라 캐시값 없음)')
    ar_counts = defaultdict(int)
    for _, n in ar_rows:
        ar_counts[n] += 1

    ambiguous = sorted(n for n in ar_counts
                       if n in amounts and len(amounts[n]) != ar_counts[n])
    matched = skipped_formula = 0
    unmatched_ar, seen = [], defaultdict(int)
    for r, n in ar_rows:
        if n in ambiguous:
            continue
        if n not in amounts:
            unmatched_ar.append(n)
            continue
        idx = seen[n]; seen[n] += 1
        b, c = amounts[n][idx]
        wrote = False
        for col, val in ((layout['detail_col'], b), (layout['main_col'], c)):
            cell = wsf.cell(row=r, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                if val is not None:
                    skipped_formula += 1
                continue
            coord = f'{get_column_letter(col)}{r}'
            old = str(wsv.cell(row=r, column=col).value)
            if val is not None:
                plan.add(op='set_cell', sheet=sheet, cell=coord, value=val, old=old)
                wrote = True
            elif cell.value is not None:
                # 원천에 없는 쪽 열의 잔존 전기값 제거 (세부/본란 위치 정합 유지)
                plan.add(op='clear_cell', sheet=sheet, cell=coord, old=old)
        if wrote:
            matched += 1
    unused_raw = sorted(n for n in amounts if n not in ar_counts)
    wbf.close(); wbv.close()

    rpt += ['', f'주입 대상: {matched}개 항목 (수식 보호로 건너뜀: {skipped_formula})',
            f'정산표에만 있고 원천에 없는 계정: {len(set(unmatched_ar))}개'
            + (' — 전기 금액이 남아 있으므로 반드시 확인' if unmatched_ar else ''),
            f'원천에만 있고 정산표에 없는 계정: {len(unused_raw)}개',
            f'출현 횟수 불일치(모호 — 주입 제외): {len(ambiguous)}개']
    if ambiguous:
        rpt.append('  모호: ' + ', '.join(ambiguous[:20]))
        plan.manual.append('출현 횟수 불일치 계정 수동 확인: ' + ', '.join(ambiguous[:20]))
    if unmatched_ar:
        rpt.append('  정산표 전용: ' + ', '.join(sorted(set(unmatched_ar))[:25]))
    if unused_raw:
        rpt.append('  원천 전용: ' + ', '.join(unused_raw[:25]))
        plan.manual.append('원천 전용 계정(신규 계정 가능) 정산표 행 추가 검토: '
                           + ', '.join(unused_raw[:15]))
    rpt.append('')
    rpt.append('※ 주입 후 정산표의 소계·합계(수식)는 Excel에서 열면 재계산됩니다. '
               '차대·합계 검증은 재계산 후 확인하십시오.')
    return plan, rpt
