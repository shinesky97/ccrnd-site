# -*- coding: utf-8 -*-
"""엑셀 산출물 이월 계획 생성 — 일반조서·계정별조서·정산표.

계획(Plan)만 만든다. 실행은 engine.py (Dry-run 확인 후).
원칙: 추측 매핑 금지 — 열 탐지 실패 시 해당 시트는 '수동 처리' 목록으로.
"""
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .engine import Plan
from .util import GISU_PAT, bump_period_label, patch_openpyxl_korean

patch_openpyxl_korean()


def _iter_cells(ws, max_row=None, max_col=None):
    for row in ws.iter_rows(min_row=1, max_row=max_row or ws.max_row,
                            max_col=max_col or ws.max_column):
        for c in row:
            yield c


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---------------------------------------------------------------- 일반조서
def plan_general_wp(src, out, prior_fye, cur_fye, keep_authors=True):
    """일반조서 당기화:
    - 전기 결산일과 같은 날짜 셀 → 당기 결산일
    - Index 시트의 그 외 날짜(작성일·검토일) → 삭제
    - 각 조서 시트에서 '작성자/검토자' 행의 날짜('일자') → 삭제
    - '제 N 기' 라벨 → 기수·연도 +1
    """
    plan = Plan(src, out)
    wb = load_workbook(src, read_only=False, data_only=False)
    for ws in wb.worksheets:
        big = ws.max_row > 2000
        scan_rows = 400 if big else None
        if big:
            plan.warnings.append(f'{ws.title}: 대용량 시트 — 상단 400행만 날짜 치환 검사')
        for c in _iter_cells(ws, max_row=scan_rows):
            v = c.value
            d = _as_date(v)
            if d is not None:
                if d == prior_fye:
                    plan.add(op='set_cell', sheet=ws.title, cell=c.coordinate,
                             value=cur_fye, old=str(d))
                elif ws.title == 'Index':
                    plan.add(op='clear_cell', sheet=ws.title, cell=c.coordinate, old=str(d))
                else:
                    # 작성자/검토자 행의 일자만 삭제
                    row_labels = [str(x.value) for x in ws[c.row] if isinstance(x.value, str)]
                    if any(('작성자' in s or '검토자' in s) for s in row_labels):
                        plan.add(op='clear_cell', sheet=ws.title, cell=c.coordinate, old=str(d))
            elif isinstance(v, str) and GISU_PAT.search(v) and len(v) <= 60:
                new, changed = bump_period_label(v)
                if changed:
                    plan.add(op='set_cell', sheet=ws.title, cell=c.coordinate, value=new, old=v)
        if not keep_authors:
            plan.warnings.append('작성자 이름 삭제 옵션은 아직 구현되지 않음 — 유지됨')
    wb.close()
    return plan


# ---------------------------------------------------------------- 계정별조서
def plan_account_wp(src, out, prior_fye, cur_fye, company):
    """계정별조서(4000) 당기화: 목차의 담당자·종료일 비움, 결산일·회사명 세팅."""
    plan = Plan(src, out)
    wb = load_workbook(src, read_only=False, data_only=False)
    toc = '입증감사절차'
    if toc in wb.sheetnames:
        ws = wb[toc]
        header_row = None
        cols = {}
        for c in _iter_cells(ws, max_row=15):
            if isinstance(c.value, str) and c.value.strip() in ('담당자', '종료일', '인덱스', '계정과목'):
                header_row = c.row
                cols[c.value.strip()] = c.column
        if header_row and '담당자' in cols:
            for r in range(header_row + 1, ws.max_row + 1):
                for key in ('담당자', '종료일'):
                    if key in cols:
                        cell = ws.cell(row=r, column=cols[key])
                        if cell.value not in (None, ''):
                            plan.add(op='clear_cell', sheet=toc,
                                     cell=cell.coordinate, old=str(cell.value)[:20])
        else:
            plan.manual.append(f'{toc}: 담당자/종료일 열 탐지 실패 — 수동 정리 필요')
    # 각 시트 상단의 결산일·회사명 라벨
    for ws in wb.worksheets:
        for c in _iter_cells(ws, max_row=8):
            d = _as_date(c.value)
            if d is not None and d == prior_fye:
                plan.add(op='set_cell', sheet=ws.title, cell=c.coordinate,
                         value=cur_fye, old=str(d))
    wb.close()
    return plan


# ---------------------------------------------------------------- 정산표
def detect_period_columns(ws, scan_rows=8):
    """헤더에서 당기/전기 열 탐지. {'cur': col, 'prv': col, 'header_row': r} 또는 None.

    판별 규칙 (탐지 실패 시 None — 추측하지 않는다):
    1) '제 N (당)기' / '제 N (전)기' 표기가 있으면 그것을 사용
    2) '당기(말)' / '전기(말)' 단어 표기
    3) '제 N 기'만 있고 (당)(전) 표기가 없으면 기수 N이 큰 쪽 = 당기
    """
    import re
    marked_cur = marked_prv = None      # (당)(전) 명시
    plain = []                          # (기수N, col, row) — 표기 없음
    hrow = None
    for row in ws.iter_rows(min_row=1, max_row=scan_rows):
        for c in row:
            if not isinstance(c.value, str):
                continue
            s = c.value.replace(' ', '').replace(' ', '')
            m = re.search(r'제(\d+)\((당|전)\)기', s)
            if m:
                if m.group(2) == '당' and marked_cur is None:
                    marked_cur, hrow = c.column, c.row
                elif m.group(2) == '전' and marked_prv is None:
                    marked_prv = c.column
                continue
            if re.search(r'^당기(말)?$', s) and marked_cur is None:
                marked_cur, hrow = c.column, c.row
                continue
            if re.search(r'^전기(말)?$', s) and marked_prv is None:
                marked_prv = c.column
                continue
            m = re.search(r'제(\d+)기', s)
            if m and '수정' not in s:
                plain.append((int(m.group(1)), c.column, c.row))
        if marked_cur and marked_prv:
            break
    if marked_cur and marked_prv:
        return {'cur': marked_cur, 'prv': marked_prv, 'header_row': hrow}
    if len(plain) >= 2:
        plain.sort(reverse=True)          # 기수 큰 순
        (n1, c1, r1), (n2, c2, _) = plain[0], plain[1]
        if n1 != n2:
            return {'cur': c1, 'prv': c2, 'header_row': r1}
    return None


def plan_trial_sheet(src, out, prior_year=None, insert_history_col=True):
    """정산표 이월 (발주자 수작업 3·4·5단계):
    3) CF·FN: 당기값 → 전기 위치로 값복사
    4) AR: 전기값 우측에 열 삽입 + 당기수치 값복사 (열 삽입은 Excel 엔진 필요)
    5) BS·PL(존재 시): 당기 → 전기 값복사
    +) 상단 헤더의 기수·연도 라벨 갱신 ('제 21(당)기'→'제 22(당)기', '2025'→'2026',
       '2024 감사후'→'2025 감사후')
    열 탐지 실패 시트는 수동 목록으로 남긴다.
    """
    plan = Plan(src, out)
    wb = load_workbook(src, read_only=True, data_only=True)

    # 헤더 라벨 갱신 (상단 10행 한정 — 데이터 영역 오염 방지)
    if prior_year:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=10):
                for c in row:
                    v = c.value
                    if isinstance(v, str) and len(v) <= 60 and (
                            GISU_PAT.search(v) or str(prior_year) in v
                            or str(prior_year - 1) in v):
                        new, changed = bump_period_label(v)
                        if changed:
                            plan.add(op='set_cell', sheet=ws.title,
                                     cell=c.coordinate, value=new, old=v)
                    elif isinstance(v, int) and v in (prior_year, prior_year - 1):
                        plan.add(op='set_cell', sheet=ws.title,
                                 cell=c.coordinate, value=v + 1, old=str(v))
                    elif isinstance(v, datetime) and v.year in (prior_year, prior_year - 1):
                        try:
                            plan.add(op='set_cell', sheet=ws.title, cell=c.coordinate,
                                     value=v.replace(year=v.year + 1), old=str(v)[:10])
                        except ValueError:   # 2/29 등
                            pass

    def copy_cur_to_prv(sheet_name):
        ws = wb[sheet_name]
        det = detect_period_columns(ws)
        if not det:
            plan.manual.append(f'{sheet_name}: 당기/전기 열 탐지 실패 — 수동 값복사 필요')
            return
        plan.add(op='note', text=f"{sheet_name}: 당기열 {get_column_letter(det['cur'])}"
                                 f" → 전기열 {get_column_letter(det['prv'])} 값복사")
        plan.add(op='copy_col_values', sheet=sheet_name,
                 src_col=det['cur'], dst_col=det['prv'],
                 row_from=det['header_row'] + 1, row_to=ws.max_row)

    for name in ('CF', 'FN', 'BS', 'PL', 'dsd'):
        if name in wb.sheetnames:
            copy_cur_to_prv(name)

    if 'AR' in wb.sheetnames:
        ws = wb['AR']
        det = detect_period_columns(ws)
        if det:
            r1, r2 = det['header_row'] + 1, ws.max_row
            # 당기 금액 2단(세부/본란) → 전기 금액 2단 값복사 (수작업 5단계)
            for off in (0, 1):
                plan.add(op='copy_col_values', sheet='AR',
                         src_col=det['cur'] + off, dst_col=det['prv'] + off,
                         row_from=r1, row_to=r2)
            plan.add(op='note', text=f"AR: 당기 금액 {get_column_letter(det['cur'])}~"
                                     f"{get_column_letter(det['cur'] + 1)} → 전기 금액 "
                                     f"{get_column_letter(det['prv'])}~"
                                     f"{get_column_letter(det['prv'] + 1)} 값복사")
            if prior_year:
                # '전기 감사후' 이력 갱신 + 당기 수정분개(차변/대변) 초기화 (수작업 4단계)
                cols = {}
                wb2 = load_workbook(src, read_only=True)   # 수식 문자열 확인용
                ws2 = wb2['AR']
                for row in ws.iter_rows(min_row=1, max_row=10):
                    for c in row:
                        if isinstance(c.value, str):
                            s = c.value.replace(' ', '')
                            if s == f'{prior_year}감사후':
                                cols['q'] = c.column
                            elif s == f'{prior_year - 1}감사후':
                                cols['m'] = c.column
                            elif s == '차변':
                                cols['dr'] = c.column
                            elif s == '대변':
                                cols['cr'] = c.column
                if 'q' in cols and 'm' in cols:
                    plan.add(op='copy_col_values', sheet='AR',
                             src_col=cols['q'], dst_col=cols['m'],
                             row_from=r1, row_to=r2)
                    plan.add(op='note', text=f"AR: {prior_year} 감사후 값"
                                             f"({get_column_letter(cols['q'])}) → 전기 감사후 "
                                             f"이력열({get_column_letter(cols['m'])}) 값복사")
                else:
                    plan.manual.append("AR: '감사후' 이력 열 탐지 실패 — 수동 갱신 필요")
                cleared = 0
                for key in ('dr', 'cr'):
                    if key in cols:
                        for r in range(r1, r2 + 1):
                            v = ws2.cell(row=r, column=cols[key]).value
                            if v is not None and not (isinstance(v, str) and v.startswith('=')):
                                plan.add(op='clear_cell', sheet='AR',
                                         cell=f"{get_column_letter(cols[key])}{r}",
                                         old=str(v)[:20])
                                cleared += 1
                if cleared:
                    plan.add(op='note', text=f'AR: 전기 수정분개(차변/대변) {cleared}셀 초기화')
                wb2.close()
        else:
            plan.manual.append('AR: 당기/전기 열 탐지 실패 — 수동 처리 필요')
    wb.close()
    plan.warnings.append('정산표 열 구조는 업체별로 다를 수 있음 — Dry-run에서 탐지 결과를 '
                         '반드시 확인 후 실행하십시오.')
    return plan
