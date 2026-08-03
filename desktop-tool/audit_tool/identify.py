# -*- coding: utf-8 -*-
"""D1: 폴더 스캔·파일 자동 식별.

파일명이 아닌 '내용'(시트 구성, DSD XML)을 근거로 분류한다. 판별이 모호하면
추측하지 않고 unknown으로 남긴다 (사용자 지정으로 해결).
"""
import os
import re
import zipfile
import xml.etree.ElementTree as ET

from .util import parse_year, GISU_PAT

KINDS = ('dsd', 'trial_sheet', 'general_wp', 'account_wp', 'raw_data')
KIND_LABEL = {
    'dsd': '전기 DSD', 'trial_sheet': '전기 정산표', 'general_wp': '전기 일반조서',
    'account_wp': '전기 계정별조서', 'raw_data': '당기 전산자료(Raw)', 'unknown': '미분류',
}


def _classify_xlsx(path):
    """시트 구성으로 엑셀 파일 종류를 판별. (읽기 전용)"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        names = set(wb.sheetnames)
        wb.close()
    except Exception as e:
        return 'unknown', {'error': f'열기 실패: {e}'}
    info = {'sheets': sorted(names)}
    if 'Index' in names and names & {'1100', '2700', '8600'}:
        return 'general_wp', info
    if '입증감사절차' in names:
        return 'account_wp', info
    if {'AR', 'SAD'} <= names:
        return 'trial_sheet', info
    if {'BS', 'PL'} <= names and (names & {'RQ', 'ARAP', '조회서', '월보'}
                                  or any(re.fullmatch(r'\d{2}', s) for s in names)):
        return 'raw_data', info
    return 'unknown', info


def _inspect_dsd(path):
    """DSD에서 회사명·기수·연도·편집기 버전을 추출."""
    info = {}
    try:
        with zipfile.ZipFile(path) as z:
            contents = z.read('contents.xml').decode('utf-8', errors='replace')
            meta = z.read('meta.xml').decode('utf-8', errors='replace')
    except Exception as e:
        return {'error': f'DSD 열기 실패: {e}'}
    m = re.search(r'regname="([^"]+)"', meta)
    if m:
        info['company'] = m.group(1)
    m = re.search(r'editver="([^"]+)"', meta)
    if m:
        info['editor_version'] = m.group(1)
    m = re.search(r'<FORMULA-VERSION ADATE="(\d+)">([\d.]+)</FORMULA-VERSION>', contents)
    if m:
        info['formula_version'] = f'{m.group(2)} ({m.group(1)})'
    m = re.search(r'제\s*(\d+)\s*\(\s*당\s*\)\s*기', contents)
    if m:
        info['fiscal_no'] = int(m.group(1))
    m = re.search(r'제\s*\d+\s*기\s*(20\d\d)년\s*\d+월\s*\d+일\s*현재', contents)
    if m:
        info['year'] = int(m.group(1))
    return info


def _detect_year_xlsx(path, kind):
    """엑셀 내부 결산일/기수 헤더에서 연도 추정 (파일명 연도는 보조)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        target = {'general_wp': 'Index', 'trial_sheet': 'AR', 'raw_data': 'BS',
                  'account_wp': '입증감사절차'}.get(kind)
        ws = wb[target] if target in wb.sheetnames else wb[wb.sheetnames[0]]
        fye_year = any_year = None
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            for v in row:
                if v is None:
                    continue
                if hasattr(v, 'year') and 2000 < v.year < 2100:
                    # 결산일 형태(12/31 등 월말·연말)를 우선, 작성일 등은 보조
                    if getattr(v, 'month', 0) == 12 and getattr(v, 'day', 0) == 31:
                        fye_year = fye_year or v.year
                    else:
                        any_year = any_year or v.year
                elif isinstance(v, str) and GISU_PAT.search(v):
                    y = parse_year(v)
                    if y:
                        fye_year = fye_year or y
            if fye_year:
                break
        wb.close()
        return fye_year or any_year
    except Exception:
        return None


def scan_folder(folder, max_depth=2):
    """폴더를 스캔해 종류별 후보 목록을 반환한다.

    반환: {'found': {kind: [ {path, year, detail}, ... ]}, 'unknown': [...]}
    """
    found = {k: [] for k in KINDS}
    unknown = []
    base_depth = folder.rstrip(os.sep).count(os.sep)
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if not d.startswith(('_audit_tool', '~', '.'))]
        if root.count(os.sep) - base_depth >= max_depth:
            dirs[:] = []
        for fn in files:
            if fn.startswith('~$'):
                continue
            path = os.path.join(root, fn)
            ext = os.path.splitext(fn)[1].lower()
            if ext == '.dsd':
                info = _inspect_dsd(path)
                found['dsd'].append({'path': path, 'year': info.get('year') or parse_year(fn),
                                     'detail': info})
            elif ext in ('.xlsx', '.xlsm'):
                kind, info = _classify_xlsx(path)
                year = _detect_year_xlsx(path, kind) or parse_year(fn)
                entry = {'path': path, 'year': year, 'detail': info}
                if kind == 'unknown':
                    unknown.append(entry)
                else:
                    found[kind].append(entry)
    return {'found': found, 'unknown': unknown}


def summarize(scan):
    lines = []
    for kind in KINDS:
        for e in scan['found'][kind]:
            lines.append(f"  [{KIND_LABEL[kind]}] {os.path.basename(e['path'])}"
                         f" (연도: {e['year'] or '판별불가'})")
    for e in scan['unknown']:
        lines.append(f"  [미분류] {os.path.basename(e['path'])}")
    return '\n'.join(lines) if lines else '  (식별된 파일 없음)'
